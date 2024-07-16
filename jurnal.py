from selenium import webdriver
from datetime import datetime
from pprint import pprint
import time
import pickle
import shutil
import os
import openpyxl

# op = webdriver.ChromeOptions()
# op.add_argument('headless')
# driver = webdriver.Chrome(options=op)

driver = webdriver.Chrome()

error_flag = False

# with open('week_data.pkl', 'rb') as file:
#     saved_data = pickle.load(file)
#
# # Access the variables from the loaded data
# week_dates = saved_data['week_dates']
# week_pars = saved_data['week_pars']
# week_types = saved_data['week_types']
# week_titles = saved_data['week_titles']
# week_times = saved_data['week_times']
# week_pars_pos = saved_data['week_pars_pos']
# week_were = saved_data['week_were']

months_dict = {'Января': 1, 'Февраля': 2, 'Марта': 3, 'Апреля': 4,
               'Мая': 5, 'Июня': 6, 'Июля': 7, 'Августа': 8,
               'Сентября': 9, 'Октября': 10, 'Ноября': 11, 'Декабря': 12}

positions = ['09:00', '10:40', '12:40', '14:20', '16:20', '18:00', '18:30']

week_num = input("Введите номер недели: ")

try:
    driver.get("https://attendance-app.mirea.ru")

    driver.implicitly_wait(10)

    button_xpath = "/html/body/div/div/a/button"
    button = driver.find_element("xpath", button_xpath)

    button_text_xpath = "/html/body/div/div/a/button/span"
    button_text = driver.find_element("xpath", button_text_xpath)

    button_format = button_text.text

    if button_format == "Войти":
        button.click()

        driver.implicitly_wait(10)

        login_xpath = "/html/body/div/div/div/div/form/div[1]/div[1]/input"
        login = driver.find_element("xpath", login_xpath)

        login.send_keys("guseynov.a.s@edu.mirea.ru")

        driver.implicitly_wait(10)

        password_xpath = "/html/body/div/div/div/div/form/div[1]/div[2]/input"
        password = driver.find_element("xpath", password_xpath)

        password.send_keys("Alex14102004")

        driver.implicitly_wait(10)

        login_button_xpath = "/html/body/div/div/div/div/form/div[2]/button"
        login_button = driver.find_element("xpath", login_button_xpath)

        login_button.click()

        driver.implicitly_wait(10)

        button_xpath = "/html/body/div/div/a/button"
        button = driver.find_element("xpath", button_xpath)

        button_text_xpath = "/html/body/div/div/a/button/span[2]"
        button_text = driver.find_element("xpath", button_text_xpath)

        button_format = button_text.text

        if button_format == "Перейти к журналу":
            button.click()

            driver.implicitly_wait(10)

            now_week_num = -1
            now_day = ""
            int_week_num = int(week_num)

            switch_left_button_xpath = "/html/body/div/div[1]/div/div/div[2]/div/div/div[1]/div/button"
            switch_left_button = driver.find_element("xpath", switch_left_button_xpath)

            switch_right_button_xpath = "/html/body/div/div[1]/div/div/div[2]/div/div/div[3]/div/button"
            switch_right_button = driver.find_element("xpath", switch_right_button_xpath)

            while now_week_num != int_week_num or now_day != "Понедельник":
                week_day_xpath = "/html/body/div/div[1]/div/div/div[2]/div/div/div[2]/div/p[2]"
                week_day = driver.find_element("xpath", week_day_xpath)

                week_day_text = week_day.text

                now_week_num = int(week_day_text.split(',')[0].split(' ')[1][1:])
                now_day = week_day_text.split(',')[1][1:]

                if now_week_num == int_week_num:
                    if now_day != "Понедельник":
                        switch_left_button.click()

                elif now_week_num < int_week_num:
                    switch_right_button.click()

                    driver.implicitly_wait(10)

                else:
                    switch_left_button.click()

                    driver.implicitly_wait(10)

            week_dates = []
            week_types = []
            week_titles = []
            week_times = []
            week_pars_pos = []
            week_pars = []
            week_were = []

            i = 1

            while i <= 6:
                try:
                    formatted_date = ''
                    types = []
                    titles = []
                    times = []
                    day_were = []
                    pars_pos = [False, False, False, False, False, False, False]

                    j = 1

                    while True:
                        try:
                            found = -1

                            date_xpath = '/html/body/div[1]/div[1]/div/div/div[2]/div/div/div[2]/div/p[1]'
                            date_text = driver.find_element("xpath", date_xpath).text

                            day, month_name, year = map(str.strip, date_text.split())
                            month = months_dict[month_name]
                            date_object = datetime(int(year[:-1]), month, int(day))
                            formatted_date = date_object.strftime('%d.%m.%Y')

                            common_element_xpath = f'/html/body/div/div[1]/div/div/div[3]/div/div[{j}]'
                            common_element = driver.find_element("xpath", common_element_xpath)

                            type_xpath = './div[1]/div/div[2]/span'
                            type_text = common_element.find_element("xpath", type_xpath).text
                            types.append(type_text)

                            title_xpath = './div[1]/div/div[1]/div/div/div[2]/div/div[1]/strong'
                            title_text = common_element.find_element("xpath", title_xpath).text
                            titles.append(title_text)

                            time_xpath = './div[1]/div/div[1]/div/div/div[1]/div[1]'
                            time_text = common_element.find_element("xpath", time_xpath).text
                            times.append(time_text)

                            for y in range(0, len(positions)):
                                if positions[y] == time_text:
                                    pars_pos[y] = True

                                    found = y

                            common_element.click()

                            time.sleep(1)

                            try:
                                common_pod_element_xpath = '/html/body/div[2]/div/div[3]/div/div/div[2]/div[2]/div[1]/div/div/div'
                                common_pod_element = driver.find_element("xpath", common_pod_element_xpath)

                                were = []

                                k = 1

                                while k <= 27:
                                    try:
                                        now_fio = common_pod_element.find_element("xpath", f'./li[{k}]')

                                        fio = now_fio.find_element("xpath", './h5').text.split()[0]

                                        P = now_fio.find_element("xpath", f'./div/label[1]').value_of_css_property("background-color")

                                        if P[16:18] == "15":
                                            was = True

                                        else:
                                            was = False

                                        were.append([fio, was])

                                    except:
                                        print("Ошибка 3", formatted_date, type_text, title_text, time_text, found)
                                        break

                                    finally:
                                        k += 1

                                day_were.append(were)

                                driver.find_element("xpath", f'/html/body/div[2]/div/div[3]/div/div/div[1]/div[1]/button/span').click()
                                driver.implicitly_wait(0.5)

                            except:
                                driver.find_element("xpath", f'/html/body/div[2]/div/div[3]/div/div/div[1]/div[1]/button/span').click()
                                driver.implicitly_wait(0.5)

                        except:
                            error_flag = True

                            break

                        finally:
                            j += 1

                    week_dates.append(formatted_date)
                    week_types.append(types)
                    week_titles.append(titles)
                    week_times.append(times)
                    week_pars_pos.append(pars_pos)
                    week_were.append(day_were)
                    week_pars.append(j - 2)

                except:
                    print('Ошибка 2', i)
                    error_flag = True

                    break

                finally:
                    i += 1

                    switch_right_button.click()

                    driver.implicitly_wait(10)

            # print('даты:\n')
            # pprint(week_dates)
            # print('\nсколько пар:\n')
            # pprint(week_pars)
            # print('\nтипы пар:\n')
            # pprint(week_types)
            # print('\nназвания пар:\n')
            # pprint(week_titles)
            # print('\nвремя начала пар:\n')
            # pprint(week_times)
            # print('\nналичие пар:\n')
            # pprint(week_pars_pos)
            # print('\nприсутствие на парах:\n')
            # pprint(week_were)

            with open('week_data.pkl', 'wb') as file:
                pickle.dump({
                    'week_dates': week_dates,
                    'week_pars': week_pars,
                    'week_types': week_types,
                    'week_titles': week_titles,
                    'week_times': week_times,
                    'week_pars_pos': week_pars_pos,
                    'week_were': week_were
                }, file)

            source_filename = 'tests.xlsx'
            new_filename = f'week_{week_num}.xlsx'
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            source_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), source_filename)
            new_path = os.path.join(desktop_path, new_filename)

            shutil.copy(source_path, new_path)

            surenames = [surename[0] for surename in week_were[3][2]]

            surenames_dict = {surname: index + 1 for index, surname in enumerate(surenames)}

            abbreviation_dict = {
                "Механика рычажных манипуляторов": "мрм",
                "Дифференциальные уравнения": "дифф ур",
                "Физика (1 п/г)": "физ",
                "Физика (2 п/г)": "физ",
                "Физика": "физ",
                "Объектно-ориентированное программирование": "ооп",
                "Физическая культура и спорт (элективная дисциплина)": "физ-ра",
                "Математический анализ": "мат ан",
                "Электротехника": "эл",
                "Иностранный язык": "ин яз",
                "Автоматизация инженерных расчетов": "аир",
                "Электротехника (1 п/г)": "эл",
                "Электротехника (2 п/г)": "эл",
                "Искусственный интеллект": "ии",
                "Теория передачи информации": "тпи"
            }

            group_one_dict = [
                'Аникина',
                'Астафьев',
                'Бадалян',
                'Виноградов',
                'Волобуев',
                'Гизетдинов',
                'Гусейнов',
                'Зимина',
                'Калашников',
                'Катасонов',
                'Коробов',
                'Куприянова',
                'Львов',
                'Лямин'
            ]

            week_podgroups = []

            for i in range(0, len(week_titles)):
                day_podgroups = []

                for j in range(0, len(week_titles[i])):
                    if '1' in week_titles[i][j]:
                        day_podgroups.append(1)

                    elif '2' in week_titles[i][j]:
                        day_podgroups.append(2)

                    else:
                        day_podgroups.append(0)

                    week_titles[i][j] = abbreviation_dict.get(week_titles[i][j])

                week_podgroups.append(day_podgroups)

            if week_titles[1] == ['физ', 'физ', 'физ', 'физ']:
                week_pars[1] = 2
                week_times[1] = week_times[1][1:3]
                week_types[1] = week_types[1][0:2]
                week_titles[1] = week_titles[1][0:2]
                week_podgroups[1] = week_podgroups[1][0:2]
                week_were[1] = week_were[1][0:2]

            for i in range(0, len(week_were)):
                for j in range(0, len(week_were[i])):
                    if week_podgroups[i][j] == 1 and week_titles[i][j] != 'физ':
                        true_flag = False

                        for k in range(0, len(week_were[i][j])):
                            if week_were[i][j][k][0] not in group_one_dict:
                                week_were[i][j][k][1] = True

                            if week_were[i][j][k][1] == True:
                                true_flag = True

                        if not true_flag:
                            week_were[i][j] = []

                    elif week_podgroups[i][j] == 2 and week_titles[i][j] != 'физ':
                        true_flag = False

                        for k in range(0, len(week_were[i][j])):
                            if week_were[i][j][k][0] in group_one_dict:
                                week_were[i][j][k][1] = True

                            if week_were[i][j][k][1] == True:
                                true_flag = True

                        if not true_flag:
                            week_were[i][j] = []

                    else:
                        true_flag = False

                        for k in range(0, len(week_were[i][j])):
                            if week_were[i][j][k][1] == True:
                                true_flag = True

                        if not true_flag:
                            week_were[i][j] = []

            wb = openpyxl.load_workbook(new_path)
            ws = wb.active

            full_week_pars_pos = []
            full_week_types = []
            full_week_titles = []

            for el_pars in week_pars_pos:
                for pod_el_pars in el_pars:
                    full_week_pars_pos.append(pod_el_pars)

            for el_types in week_types:
                for pod_el_types in el_types:
                    full_week_types.append(pod_el_types.lower())

            for el_titles in week_titles:
                for pod_el_titles in el_titles:
                    full_week_titles.append(pod_el_titles)

            try:
                wrote_cols = []

                i = 0
                j = 0
                k = 0

                for col_num in range(3, 45):
                    if full_week_pars_pos[col_num - 3]:
                        ws.cell(row=5, column=col_num, value=full_week_types[i])

                        i += 1

                        ws.cell(row=7, column=col_num, value=full_week_titles[k])

                        k += 1

                        wrote_cols.append(col_num)

                    if col_num % 7 == 3:
                        ws.cell(row=6, column=col_num, value=week_dates[j])

                        j += 1

                num_of_week_pars = len(full_week_types)
                ws.cell(row=7, column=45, value=num_of_week_pars)

                real_surenames_dict = {}
                surenames_n_dict = {}

                for i in range(9, len(surenames_dict) + 9):
                    fio = ws.cell(row=i, column=2).value.split()[0]
                    real_surenames_dict[fio] = i - 8
                    surenames_n_dict[fio] = 0

                h = 0

                for i in range(0, len(week_were)):
                    for j in range(0, len(week_were[i])):
                        for k in range(0, len(week_were[i][j])):
                            if not week_were[i][j][k][1]:
                                ws.cell(row=real_surenames_dict[week_were[i][j][k][0]] + 8, column=wrote_cols[h],
                                        value='н')

                                surenames_n_dict[week_were[i][j][k][0]] += 1

                        h += 1

                m = 9

                for key, value in surenames_n_dict.items():
                    ws.cell(row=m, column=45, value=num_of_week_pars)

                    ws.cell(row=m, column=46, value=value)

                    m += 1


            except:
                print("error")

            wb.save(new_path)
            wb.close()

            print("Таблица готова")

        else:
            error_flag = True

    else:
        error_flag = True

finally:
    # if not error_flag:
    #     print("Готово")
    #
    # else:
    #     print("Ошибка 1")

    driver.quit()
